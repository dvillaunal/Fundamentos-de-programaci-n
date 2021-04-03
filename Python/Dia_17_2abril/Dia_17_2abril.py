# Manejo de archivos de Excel en Python

'''
Los mismos principios de manipulación de archivos de texto
aplican para hojas de excel, con algunas diferencias:


-> Crear archivo excel en python:

Como ya sabrás, en terminología de hojas de cálculo, las hojas de cálculo se agrupan en libros.
Un libro es la entidad superior de una hoja de cálculo (se suele corresponder con un archivo excel).
A su vez, un libro está compuesto por una o más hojas. Un libro siempre está formado,
al menos, por una hoja y la hoja en la que se está trabajando se denomina hoja activa.

Para comenzar a trabajar con openpyxl,
no hace falta guardar ningún archivo en el sistema de ficheros.
Simplemente hay que crear un libro.
'''
import openpyxl
import csv
from io import open

wb = openpyxl.Workbook()

ha = wb.active
ha.title = 'Ensayo 00'
print("Hoja activa => {}" .format(ha.title))

'''
Crear una hoja:

Además de la hoja por defecto,
con openpyxl es posible crear más hojas en un libro
utilizando el método create_sheet() del workbook:
'''

hoja1 = wb.create_sheet('Hoja')

hoja2 = wb.create_sheet('Hoja', 0)

print(ha.title)

'Nota: También es posible crear una copia de una hoja con el método copy_worksheet():'

o = wb.active
copy = wb.copy_worksheet(o)

'''
Acceder a una hoja:

los nombres de las hojas son una propiedad muy importante
ya que nos permiten acceder a ellas directamente
tratando el workbook como un diccionario. 
'''

ha = wb.active
ha.title = 'Hoja1'
print("Hoja activa => {}" .format(ha.title))

ha = wb.active
ha.title = 'Hoja'
print("Hoja activa => {}" .format(ha.title))

print('-------------------------Nombres de las Hojas-------------------------')
list_names = wb.sheetnames
for h in range(len(list_names)):
  print(list_names[h])

'''
Acceder a una celda

Hasta aquí hemos visto cómo crear un libro,
las hojas y cómo acceder a ellas. Ahora vamos a lo importante,
cómo acceder al valor de una celda y cómo guardar un dato.

Es posible acceder a una celda tratando a la hoja como un diccionario,
de tal manera que se usa como clave el nombre de la celda.
Este resulta de la combinación del nombre de la columna y el número de la fila.

celda de la columna A y la fila 1: 'A1'
'''

a8 = ha['A8']
print(a8.value)

'Nota: También es posible acceder a una celda usando la notación fila, columna con el método cell()'

c8 = ha.cell(row=8, column=3)
print(c8.value)



Nota_I = open('Importante.txt', 'w')

Nota_I.write('IMPORTANTE:\nCuando se crea un libro, este no contiene ninguna celda.\nLas celdas se van creando en memoria a medida que se va accediendo a ellas, aunque no contengan ningún valor.')

Nota_I.close()

Nota_I = open('Importante.txt', 'r')

nota = Nota_I.read()

Nota_I.close()

print('-------------------------Ojo Con la Siguiente Nota-------------------------')
print(nota)
print('-------------------------Fin-------------------------')

'''
Escribir valores en una celda
En la sección anterior habrás podido comprobar que al mostrar
el contenido de una celda (print(a1.value)) siempre devolvía None.
Esto se debe a que la celda no contiene ningún valor.

Para asignar un valor a una celda en concreto, se puede hacer de tres formas diferentes:
'''
# 1.- Asiganado el valor directamente a la celda:
ha['A1'] = 'Celda A1'
a1 = ha['A1']
print(a1.value)

ha['A8'] = 81
a8 = ha['A8']
print(a8.value)

# 2.- Usando la notacione, fila columna, con el argumento value:
c1 = ha.cell(row=1, column=3, value='2do metodo')
print(c1.value)

c3 = ha.cell(row=3, column=3, value=62)
print(c3.value)

# 3.- Actualizando la propiedad value de una celda:
c8 = ha.cell(row=8, column=3)
c8.value = 2341
print(c8.value)

'''
Guardar una lista de valores:

Asignar un valor a una celda puede estar bien para determinados casos.
Sin embargo, en Python suele ocurrir que los datos están almacenados en listas o tuplas.
Para estos casos, en los que se requiere exportar determinados datos
'''
Fruver = [
          ('Manzana Roja', 'mr132', 10000, 1000),
          ('Manzana Verde', 'mv139', 10000, 1000),
          ('Pera', 'p001', 8000, 800),
          ('Melón', 'm010', 15000, None),
          ('Sandia', 's032', 12000, 2000),
          ('Fresa', 'f234', 5000, 200),
          ('Maracuya', 'ma216', 4000, 1500),
          ('Banano', 'b210', 3000, 300),
]

T_Fruver = ('Nombre', 'Referencia', 'Precio/KG', 'Precio/u.')


frutas = openpyxl.Workbook()
fhoja = frutas.active

fhoja.append(T_Fruver)

for f in Fruver:
  fhoja.append(f)

'''
Guardar un libro excel en Python:

Para guardar un archivo excel usando openpyxl,
solo tienes que llamar al método save() del workbook indicando el nombre del archivo.
Esto guardará el libro con todas las hojas y los datos de cada una de ellas.
'''


frutas.save('Productos.xlsx')
wb.save('Mi_Hoja_Cal.xlsx')

# Mejor visualización:
'''
Profe para que pueda visualizar mejor el archivo sin descargarlo y le de más seguridad,
lo converti a .csv, para que pueda verlo desde la misma app web
'''

print('Profe para que pueda visualizar mejor el archivo sin descargarlo y le de más seguridad, lo converti a .csv, para que pueda verlo desde la misma app web')

data_fhoja = fhoja.rows
data_ha = ha.rows


csv_fhoja = open("data_frutas.csv", "w+")
csv_ha = open("data_ha.csv", "w+")

# data_frutas:
for row in data_fhoja:
    l = list(row)
    for i in range(len(l)):
        if i == len(l) - 1:
            csv_fhoja.write(str(l[i].value))
        else:
            csv_fhoja.write(str(l[i].value) + ',')
        csv_fhoja.write('\n')

# data_ha:
for row in data_ha:
    l = list(row)
    for i in range(len(l)):
        if i == len(l) - 1:
            csv_ha.write(str(l[i].value))
        else:
            csv_ha.write(str(l[i].value) + ',')
        csv_ha.write('\n')

csv_fhoja.close()
csv_ha.close()
