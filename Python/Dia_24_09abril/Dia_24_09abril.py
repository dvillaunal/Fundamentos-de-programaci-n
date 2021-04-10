import openpyxl
import numpy as np
e1 = openpyxl.load_workbook('Ejercicio1.xlsx')
# Ejercicio 1:
'''
Comprobar si cada elemento de un arreglo NumPy es:
  -> un real o
  -> es un escalar.
'''
h = e1['Hoja0']

l_int=[]
for row in h.iter_rows():
    p = row[0].value
    l_int.append(p)

a_int = np.array(l_int)


l_float=[]
for row in h.iter_rows():
    p = row[1].value
    l_float.append(p)

a_float = np.array(l_float)

'Crea otro Array donde muestra True si cumple y false en caso contrario'

ai_bool = np.isreal(a_int)

af_bool = np.iscomplex(a_float)

# Resultados 1:
wb = openpyxl.Workbook()
hoja = wb.active
hoja.title = 'R1'

hoja.cell(row=1, column=1, value='True: 1 & FALSE: 0')
hoja.cell(row=1, column=4, value='Comprobar si cada elemento de un arreglo NumPy es: Real o Complejo')
f = 2
col_ai = 1
col_af = 2

for x, y in zip(ai_bool, af_bool):
  hoja.cell(column= col_ai, row= f, value=x)
  hoja.cell(column= col_af, row= f, value=y)
  f += 1

wb.save(filename = 'Resultados1.xlsx')

# Ejercicio 2:
'Calcular el seno, coseno y tangente de un arreglo (valores en radianes (Decimal))'
e2 = openpyxl.load_workbook('Ejercicio2.xlsx')

h = e2['Hoja1']

l = []
for row in h.iter_rows():
    p = row[0].value
    l.append(p)

sin = np.sin(l)
cos = np.cos(l)
tan = np.tan(l)

# Resultados 2:
'Calcular Seno, Coseno y Tangante para un Arreglo'
wb = openpyxl.Workbook()
hoja = wb.active
hoja.title = 'R2'

hoja.cell(row=1, column=1, value='[PI/6, PI/4, PI/3, PI/2]')
hoja.cell(row=1, column=4, value='Calcular el seno, coseno y tangente de un arreglo (valores en radianes)')
hoja.cell(row=2, column=1, value='Sen(x)')
hoja.cell(row=2, column=2, value='Cos(x)')
hoja.cell(row=2, column=3, value='Tan(x)')
f = 3
col_s = 1
col_c = 2
col_t = 3

for x, y, z in zip(sin,cos,tan):
  hoja.cell(column= col_s, row= f, value=x)
  hoja.cell(column= col_c, row= f, value=y)
  hoja.cell(column= col_t, row= f, value=z)
  f += 1

wb.save(filename = 'Resultados2.xlsx')


# Ejercicio 3:
'Dividir Cada Fila de una Matriz por un Arreglo'
e3 = openpyxl.load_workbook('Ejercicio3.xlsx')

h = e3['Hoja1']

l0 = []
for col in h.iter_cols():
    p = col[0].value
    l0.append(p)

l1 = []
for col in h.iter_cols():
    p = col[1].value
    l1.append(p)

l2 = []
for col in h.iter_cols():
    p = col[2].value
    l2.append(p)

m = np.array([l0,l1,l2])


a = np.array(l0)

div = m/a

# Resultados 3:
'Dividir Cada Fila de una Matriz por un Arreglo'
wb = openpyxl.Workbook()
hoja = wb.active
hoja.title = 'R3'

hoja.cell(row=1, column=1, value='Dividir Cada Fila de una Matriz por un Arreglo')
f = 2
col_1 = 1
col_2 = 2
col_3 = 3

for x, y, z in zip(div[0], div[1], div[2]):
  hoja.cell(column= col_1, row= f, value=x)
  hoja.cell(column= col_2, row= f, value=y)
  hoja.cell(column= col_3, row= f, value=z)
  f += 1

wb.save(filename = 'Resultados3.xlsx')
